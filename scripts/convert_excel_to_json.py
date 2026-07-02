import json
import os
from openpyxl import load_workbook

SOURCE = r'c:\Users\samsung\Videos\dindin goumet.xlsx'
OUTPUT = r'c:\Users\samsung\Documents\CRM CODE\public\dindin-gourmet-data.json'


def normalize_text(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def parse_sheet_rows(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([normalize_text(cell) for cell in row])
    return rows


def build_structure():
    wb = load_workbook(SOURCE, data_only=True)
    data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = parse_sheet_rows(ws)
        data[sheet_name] = rows

    return data


if __name__ == '__main__':
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    structure = build_structure()
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(structure, f, ensure_ascii=False, indent=2)
    print(f'Wrote {OUTPUT}')
